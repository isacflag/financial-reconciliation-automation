from openpyxl.styles import PatternFill


VERSION = "1.0.0"

AUDIT_HEADERS = (
    "CRITÉRIO",
    "CONFIANÇA",
    "PÁGINA PDF",
    "CREDOR ERP",
    "DOCUMENTO ERP",
    "LANÇAMENTO ERP",
    "VENCIMENTO ERP",
    "VALOR ERP",
)

FILL_HIGH = PatternFill("solid", fgColor="C6EFCE")
FILL_MEDIUM = PatternFill("solid", fgColor="FFEB9C")
FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")

# Keep aliases generic in the public repository. Organization-specific aliases
# belong in a private configuration file, not in source control.
COST_CENTER_ALIASES: dict[str, str] = {}
