from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from .excel_io import (
    add_audit_columns,
    find_column,
    find_header,
    read_dda_records,
    write_match,
    write_unmatched,
)
from .matching import match_records
from .models import MatchingPolicy, ProcessSummary
from .pdf_io import parse_pdf_tables


def _validate_paths(dda_path: Path, pdf_path: Path, output_path: Path) -> None:
    if not dda_path.is_file():
        raise FileNotFoundError(f"DDA spreadsheet not found: {dda_path}")
    if not pdf_path.is_file():
        raise FileNotFoundError(f"ERP PDF not found: {pdf_path}")
    if dda_path.suffix.lower() != ".xlsx":
        raise ValueError("The DDA input must be an .xlsx file")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError("The ERP input must be a .pdf file")
    if output_path.resolve() == dda_path.resolve():
        raise ValueError("Output path must not overwrite the source spreadsheet")


def process(
    dda_path: Path,
    pdf_path: Path,
    output_path: Path,
    policy: MatchingPolicy | None = None,
) -> ProcessSummary:
    _validate_paths(dda_path, pdf_path, output_path)
    effective_policy = policy or MatchingPolicy()
    pdf_records = parse_pdf_tables(pdf_path)

    workbook = load_workbook(dda_path)
    sheet = workbook.active
    header_row = find_header(sheet)
    columns = {
        "bank": find_column(sheet, header_row, ("BANCO",)),
        "beneficiary": find_column(
            sheet, header_row, ("NOME", "BENEFICIÁRIO")
        ),
        "document": find_column(
            sheet, header_row, ("Nº DOC", "N° DOC", "DOCUMENTO", "DOC")
        ),
        "date": find_column(sheet, header_row, ("VENCIMENTO",)),
        "value": find_column(sheet, header_row, ("A PAGAR", "PAGAR")),
        "cost_center": find_column(sheet, header_row, ("CENTRO DE CUSTO",)),
        "status": find_column(sheet, header_row, ("LANÇADO", "STATUS")),
    }

    dda_records = read_dda_records(sheet, header_row, columns)
    matches, candidates_by_dda, used_pdf_ids = match_records(
        dda_records, pdf_records, effective_policy
    )
    audit_columns = add_audit_columns(sheet, header_row, columns["status"])
    pdf_start = min(record.due_date for record in pdf_records)
    pdf_end = max(record.due_date for record in pdf_records)
    pdf_keys = {(record.due_date, record.cents) for record in pdf_records}

    counters = Counter()
    for dda in dda_records:
        candidate = matches.get(dda.row)
        if candidate:
            write_match(sheet, dda, candidate, columns, audit_columns)
            counters["matched"] += 1
            if candidate.confidence == "MÉDIA":
                counters["review_required"] += 1
            else:
                counters["auto_matched"] += 1
                counters["ppc" if candidate.pdf.ppc else "launched"] += 1
            continue

        if dda.due_date < pdf_start or dda.due_date > pdf_end:
            reason = (
                "FORA DO PERÍODO DO PDF "
                f"({pdf_start:%d/%m/%Y} A {pdf_end:%d/%m/%Y})"
            )
            counters["outside_period"] += 1
        elif candidates_by_dda.get(dda.row) or (
            dda.due_date,
            dda.cents,
        ) in pdf_keys:
            reason = "CONFLITO OU REGISTRO DO PDF JÁ UTILIZADO"
            counters["conflicts"] += 1
        else:
            reason = "NÃO LOCALIZADO NO PDF"
            counters["not_found"] += 1
        write_unmatched(sheet, dda, reason, columns, audit_columns)

    sheet.freeze_panes = sheet.cell(header_row + 1, 1).coordinate
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return ProcessSummary(
        dda_rows=len(dda_records),
        pdf_records=len(pdf_records),
        matched=counters["matched"],
        auto_matched=counters["auto_matched"],
        review_required=counters["review_required"],
        launched=counters["launched"],
        ppc=counters["ppc"],
        not_found=counters["not_found"],
        outside_period=counters["outside_period"],
        conflicts=counters["conflicts"],
        unused_pdf_records=len(pdf_records) - len(used_pdf_ids),
        output_path=output_path,
    )


def summary_text(summary: ProcessSummary) -> str:
    return (
        "Financial reconciliation completed.\n\n"
        f"DDA rows analyzed: {summary.dda_rows}\n"
        f"ERP records read: {summary.pdf_records}\n"
        f"Matched candidates: {summary.matched}\n"
        f"  - Automatically matched: {summary.auto_matched}\n"
        f"  - Manual review required: {summary.review_required}\n"
        f"  - LAUNCHED: {summary.launched}\n"
        f"  - PPC: {summary.ppc}\n"
        f"Not found: {summary.not_found}\n"
        f"Outside PDF period: {summary.outside_period}\n"
        f"Blocked conflicts: {summary.conflicts}\n"
        f"Unused ERP records: {summary.unused_pdf_records}\n\n"
        f"Output: {summary.output_path}"
    )
