import datetime as dt
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from financial_reconciliation.models import PdfRecord
from financial_reconciliation.service import process


class WorkflowTests(unittest.TestCase):
    def test_process_creates_auditable_output(self):
        base = dt.date(2026, 8, 10)
        pdf_records = [
            PdfRecord(
                0,
                1,
                "PROJECT ALPHA",
                "ACME LTDA",
                "NF 123456",
                "L-001",
                base,
                10000,
                False,
            ),
            PdfRecord(
                1,
                1,
                "PROJECT BETA",
                "BETA SERVICES",
                "NF 888888",
                "L-002",
                base,
                20000,
                True,
            ),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            dda_path = temp / "dda.xlsx"
            pdf_path = temp / "report.pdf"
            out_path = temp / "output.xlsx"
            pdf_path.write_bytes(b"test placeholder")

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "BANCO",
                    "NOME",
                    "Nº DOC",
                    "VENCIMENTO",
                    "A PAGAR",
                    "CENTRO DE CUSTO",
                    "LANÇADO",
                ]
            )
            sheet.append(["001", "ACME", "123456", base, 100.01, "", ""])
            sheet.append(["001", "UNKNOWN", "", base, 200.00, "", ""])
            workbook.save(dda_path)

            with patch(
                "financial_reconciliation.service.parse_pdf_tables",
                return_value=pdf_records,
            ):
                summary = process(dda_path, pdf_path, out_path)

            result = load_workbook(out_path, data_only=True).active
            headers = [
                result.cell(1, column).value
                for column in range(1, result.max_column + 1)
            ]
            self.assertEqual(summary.matched, 2)
            self.assertEqual(summary.auto_matched, 1)
            self.assertEqual(summary.review_required, 1)
            self.assertEqual(result["G2"].value, "LANÇADO")
            self.assertEqual(result["G3"].value, "REVISAR")
            self.assertIn("CRITÉRIO", headers)
            self.assertIn("CONFIANÇA", headers)


if __name__ == "__main__":
    unittest.main()
